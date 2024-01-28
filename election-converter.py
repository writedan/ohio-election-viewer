# This file converts the workbooks provided by the Ohio Secretary of State from preinct-level results into aggregates municipality-level results and pushes them into the database.
# Election workbooks are not included with this repository.

# Dependencies: pandas, openpyxl

import pandas as pd
import openpyxl
import sqlite3
import sys
import re
import os

def find_matching_files(directory_path, pattern):
    matching_files = []

    regex_pattern = re.compile(pattern)

    for filename in os.listdir(directory_path):
        if regex_pattern.match(filename):
            matching_files.append(filename)

    return matching_files

if len(sys.argv) < 3:
	print("Usage: election-converter.py <election year> <election type>")
	exit(0)

electionYear = int(sys.argv[1])
electionType = sys.argv[2]

database = sqlite3.connect('elections.db')
cursor = database.cursor()


workbookUri = 'elections/' + str(electionYear) + '/'

subdivisionWorkbookUri = workbookUri + electionType + '-subdivision-codes.xlsx'
print("Load:", subdivisionWorkbookUri)
subdivisionWorkbook = openpyxl.load_workbook(subdivisionWorkbookUri)

precinctWorkbookUri = workbookUri + electionType + '-precinct-conversion.xlsx'
print("Load:", precinctWorkbookUri)
precinctWorkbook = openpyxl.load_workbook(precinctWorkbookUri)

# electionWorkbookUri = 'elections/' + str(electionYear) + '/' + electionType + '-election.xlsx'
# print("Load:", electionWorkbookUri)
# electionWorkbook = openpyxl.load_workbook(electionWorkbookUri)

electionWorkbooks = find_matching_files(workbookUri, r"\w+-election(?:-\d+)?\.xlsx")
if len(electionWorkbooks) == 0:
	print("There must be at least one election workbook.")
	print("Election workbooks must be named as {election type}-election.xlsx or {election type}-election-{number}.xlsx")
	exit(0)
else:
	print("Found",len(electionWorkbooks),"election workbooks.")

# electionWorkbook = workbookUri + electionWorkbooks[0]
# print("Load:", electionWorkbook)
# electionWorkbook = openpyxl.load_workbook(electionWorkbook)

# # First, we will extract the election date and fill in the election data
# electionDate = electionWorkbook['Contents']['A1'].value.split(', ')[0]
# electionName = electionWorkbook['Contents']['A1'].value.split(', ')[1].split('\n')[0]

# electionWorkbook.close()

electionDate = precinctWorkbook.sheetnames[0].split(', ')[0]
electionName = precinctWorkbook.sheetnames[0].split(', ')[1].split('\n')[0]

print(f"Adding to election index: {electionName}")

cursor.execute(f"INSERT INTO election_info(name, date, year, type) VALUES(?, ?, ?, ?)", (electionName, electionDate, electionYear, electionType))

# we need to get our new id
electionInfoId = cursor.lastrowid


# Now we can begin adding counties and municipalities
data = subdivisionWorkbook[subdivisionWorkbook.sheetnames[0]].values
columns = next(data) # isolate the first row as columns
df = pd.DataFrame(data, columns=columns)
countyIndex = {}

for idx, some in df.iterrows():
	countyIndex[some['COUNTYNAME']] = some['COUNTYFP']

for idx in countyIndex:
	name = idx
	fips = countyIndex[idx]

	print("Processing county", name)

	cursor.execute(f"INSERT INTO county(name, fips, electionId) VALUES(?, ?, ?)", (name, fips, electionInfoId))

	countyId = cursor.lastrowid

	# now here we must begin adding municipalities
	muns = df[df['COUNTYFP'] == fips]
	for idx, m in muns.iterrows():
		mName = m['COUSUBNAME']
		mCode = int(m['COUSUBFP'])

		print("\tProcessing municipality", mName)
		
		cursor.execute(f"INSERT INTO municipality(name, fips, countyId) VALUES(?, ?, ?)", (mName, mCode, countyId))

# now we can setup the precincts with their conversions
data = precinctWorkbook[precinctWorkbook.sheetnames[0]].values
columns = next(data)
df = pd.DataFrame(data, columns=columns)

for idx, precinct in df.iterrows():
	countyName = precinct['COUNTYNAME'] + ' County'
	municipalCode = int(precinct['MUNICIPALFIPS'])
	precinctName = precinct['PRECINCTNAME']

	print("Processing precinct", precinctName,"of",countyName)

	cursor.execute(f"SELECT * FROM municipalities WHERE electionId=? AND countyName=? AND municipalCode=?", (electionInfoId, countyName, municipalCode))

	mRes=cursor.fetchall()
	if len(mRes) == 0:
		print("\tERROR!", countyName, municipalCode,"not in database!")
		print("\t...ocurred on", precinctName)
		print("\t...We can recover but it is best to fix this.")
		exit()
	for m in mRes:
		cursor.execute(f"INSERT INTO precinct(name, municipalId) VALUES(?, ?)", (precinctName, m[0]))

cursor.execute("create table precincts_idx as select * from precincts")
		# this will dramatically speed up the insertation process

def processElectionWorkbook(electionWorkbook):
	for worksheet in electionWorkbook.sheetnames:
		if (worksheet == 'Contents' or worksheet == 'Master'):
			continue

		cursor.execute(f"INSERT INTO office_category(name, electionId) VALUES (?, ?)", (worksheet, electionInfoId))
		officeCategoryId = cursor.lastrowid

		print("Processing election category", worksheet)

		worksheet = electionWorkbook[worksheet]

		# offices = {}
		lastOffice = ''
		officeElectionId = -1
		columnsToReckon = []
		candidatesToReckon = []

		maxColumnCount = worksheet.max_column

		for columnIdx in range(9, worksheet.max_column + 2): # 9th column idx = 'I'
			val = worksheet.cell(row=1, column=columnIdx).value
			if (val is not None) or (columnIdx > maxColumnCount):
				# first, we must handle the candidates already found
				if len(columnsToReckon) > 0:
					print("\t\tProcessing precinct results...")

					precinctId = -1
					for rowIdx in range(5, worksheet.max_row):
						countyName = worksheet.cell(row=rowIdx, column=1).value + ' County'
						precinctName = worksheet.cell(row=rowIdx, column=2).value
						cursor.execute(f"SELECT id FROM precincts_idx WHERE countyName=? AND precinctName=? AND electionId=?", (countyName, precinctName, electionInfoId))
						# print(countyName, precinctName)
						pRes = cursor.fetchall()
						if len(pRes) == 0:
							print("\t\t\tWARN!",countyName,precinctName,"not in database!")
							continue
						precinctId = pRes.pop()[0]

						if len(columnsToReckon) is not len(candidatesToReckon):
							print("\t\t\tWARN!",len(columnsToReckon),"candidates identified but only",len(candidatesToReckon),"present!")
							print("\t\t\t...ocurred at ",countyName,precinctName)

						if all(worksheet.cell(row=rowIdx, column=column).value == 0 for column in columnsToReckon):
							continue

						for idx, column in enumerate(columnsToReckon):
							candidateVotes = worksheet.cell(row=rowIdx, column=column).value
							candidateName = worksheet.cell(row=2,column=column).value
							candidateId = candidatesToReckon[idx]
							cursor.execute(f"INSERT INTO office_result(votes, candidateId, precinctId) VALUES(?, ?, ?)", (candidateVotes, candidateId, precinctId))
								
			if (val is not None):
				val = val.strip().replace('\n', ' - ')
				print("\tProcessing office", val)
				# we have a new office
				lastOffice = val
				columnsToReckon = []
				candidatesToReckon = []
				cursor.execute(f"INSERT INTO office_election(name, categoryId) VALUES(?, ?)", (lastOffice, officeCategoryId))
				officeElectionId = cursor.lastrowid

			candidateName = worksheet.cell(row=2, column=columnIdx).value
			if candidateName is None:
				continue

			candidateName.replace('\n', ' - ')
			print("\t\tProcessing candidate", candidateName)

			if (candidateName[-1] == '*'):
				print("\t\t\tWrite-in candidates are not collated. Rejected.")
				continue

			cursor.execute(f"INSERT INTO candidate(name, officeId) VALUES(?, ?)", (candidateName, officeElectionId))
			candidatesToReckon.append(cursor.lastrowid)
			columnsToReckon.append(columnIdx)

for idx, electionWorkbook in enumerate(electionWorkbooks):
	print("Load:",electionWorkbook)
	electionWorkbook = openpyxl.load_workbook(workbookUri + electionWorkbook)
	processElectionWorkbook(electionWorkbook)
	electionWorkbook.close()

cursor.execute("drop table precincts_idx") # but dont let it linger